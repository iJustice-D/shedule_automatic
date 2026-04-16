from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlmodel import Session

from app.core.timetable import DAYS, PAIR_NUMBERS
from app.services.exporters.context import build_schedule_context


class XlsxExporter:
    def export(self, session: Session, schedule_id: int, output_path: Path) -> Path:
        context = build_schedule_context(session, schedule_id)
        workbook = Workbook()
        workbook.remove(workbook.active)
        for title, grid in context["group_grids"].items():
            self._build_sheet(workbook, title[:31], grid, context, context["group_online_rows"].get(title, []))
        for title, grid in context["teacher_grids"].items():
            self._build_sheet(workbook, f"П-{title}"[:31], grid, context, context["teacher_online_rows"].get(title, []))
        self._build_balance_sheet(workbook, context)
        self._build_unresolved_sheet(workbook, context)
        workbook.save(output_path)
        return output_path

    @staticmethod
    def _build_sheet(
        workbook: Workbook,
        title: str,
        grid: dict[tuple[int, int], str],
        context: dict,
        online_rows: list[dict[str, str]],
    ) -> None:
        sheet = workbook.create_sheet(title)
        header_fill = PatternFill("solid", fgColor="E6B800")
        body_fill = PatternFill("solid", fgColor="F6F1DD")
        sheet.cell(1, 1, "Основное расписание").font = Font(bold=True, size=14)
        sheet.cell(2, 1, "День / Пара").fill = header_fill
        for column, pair_number in enumerate(PAIR_NUMBERS, start=2):
            cell = sheet.cell(2, column, context["pair_headers"][pair_number])
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical="center")
        for row, (day_of_week, _day_label) in enumerate(DAYS.items(), start=3):
            day_cell = sheet.cell(row, 1, context["day_labels"][day_of_week])
            day_cell.fill = header_fill
            day_cell.font = Font(bold=True)
            for column, pair_number in enumerate(PAIR_NUMBERS, start=2):
                cell = sheet.cell(row, column, grid.get((day_of_week, pair_number), ""))
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.fill = body_fill
        online_start = len(DAYS) + 5
        sheet.cell(online_start, 1, "Дополнительные онлайн-занятия").font = Font(bold=True, size=14)
        headers = ["День", "Онлайн-слот", "Предмет", "Преподаватель / Группа", "Формат", "Недели"]
        for column, label in enumerate(headers, start=1):
            cell = sheet.cell(online_start + 1, column, label)
            cell.fill = header_fill
            cell.font = Font(bold=True)
        if online_rows:
            for row, item in enumerate(online_rows, start=online_start + 2):
                sheet.cell(row, 1, item.get("day", ""))
                sheet.cell(row, 2, item.get("online_slot", ""))
                sheet.cell(row, 3, item.get("subject", ""))
                sheet.cell(row, 4, item.get("teacher") or item.get("group", ""))
                sheet.cell(row, 5, item.get("format", ""))
                sheet.cell(row, 6, item.get("weeks", ""))
        else:
            sheet.cell(online_start + 2, 1, "Онлайн-занятия не назначены")
        sheet.freeze_panes = "B3"
        for column in "ABCDEFG":
            sheet.column_dimensions[column].width = 26

    @staticmethod
    def _build_balance_sheet(workbook: Workbook, context: dict) -> None:
        sheet = workbook.create_sheet("Баланс нагрузки")
        header_fill = PatternFill("solid", fgColor="E6B800")
        rows = context.get("teacher_balance_rows", [])
        headers = ["Преподаватель", "Семестр 3", "Семестр 4", "Отклонение", "На уточнении"]
        for column, label in enumerate(headers, start=1):
            cell = sheet.cell(1, column, label)
            cell.fill = header_fill
            cell.font = Font(bold=True)
        if not rows:
            sheet.cell(2, 1, "Данные отсутствуют")
            return
        for row_index, row in enumerate(rows, start=2):
            sheet.cell(row_index, 1, row.get("teacher_name", ""))
            sheet.cell(row_index, 2, row.get("semester_3_pairs", ""))
            sheet.cell(row_index, 3, row.get("semester_4_pairs", ""))
            sheet.cell(row_index, 4, row.get("normalized_balance_score", ""))
            sheet.cell(row_index, 5, row.get("pending_rows", ""))

    @staticmethod
    def _build_unresolved_sheet(workbook: Workbook, context: dict) -> None:
        sheet = workbook.create_sheet("Вакансии")
        header_fill = PatternFill("solid", fgColor="E6B800")
        rows = context.get("unresolved_weekly_rows_report", [])
        headers = ["Группа", "Семестр", "Предмет", "Подгруппа", "Состояние", "Преподаватели из источника"]
        for column, label in enumerate(headers, start=1):
            cell = sheet.cell(1, column, label)
            cell.fill = header_fill
            cell.font = Font(bold=True)
        if not rows:
            sheet.cell(2, 1, "Неразрешённых строк нет")
            return
        for row_index, row in enumerate(rows, start=2):
            sheet.cell(row_index, 1, row.get("group", ""))
            sheet.cell(row_index, 2, row.get("semester", ""))
            sheet.cell(row_index, 3, row.get("subject", ""))
            sheet.cell(row_index, 4, row.get("subgroup", ""))
            sheet.cell(row_index, 5, row.get("assignment_state", ""))
            sheet.cell(row_index, 6, row.get("teacher_names", ""))
